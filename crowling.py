from selenium.webdriver.chrome.service import Service
from selenium import webdriver
import re 
import requests 
from selenium.webdriver.common.by import By
from urllib.request import urlopen 
from bs4 import BeautifulSoup
import time
import pandas as pd
import json
from openpyxl import Workbook, load_workbook


# 크롬 드라이버 경로 및 URL
driver_path = "C:/Users/tlsgo/Downloads/chromedriver-win64/chromedriver-win64/chromedriver.exe"
URL = 'https://www.q-net.or.kr/crf005.do?id=crf00501&gSite=Q&gId=#none'

# 크롬 드라이버 사용
service = Service(driver_path)

# Chrome WebDriver 실행
driver = webdriver.Chrome(service=service)

# 크롬 실행
try:
    driver.get(url=URL)
    time.sleep(3)  # 페이지 로딩을 기다리기 위한 대기 시간
    print(f"[INFO] 페이지 로딩 완료")
except Exception as e:
    print(f"오류 발생: {e}")

# 직무 분야 정보 추출
soup = BeautifulSoup(driver.page_source, 'html.parser')

# 카테고리 결과를 저장할 리스트
job_areas = []

# b 태그 중 class="tit"인 태그 찾기
b_tag = soup.find('b', class_='tit')

# b 태그를 찾았을 경우, 그 아래의 ul에서 li 태그들을 찾고, 각 li의 span 값을 추출
if b_tag:
    # b 태그 아래의 ul 찾기
    ul_tag = b_tag.find_next('ul')
    if ul_tag:
        # ul 안의 모든 li 태그에서 span 값을 추출
        spans = ul_tag.find_all('li')
        for li in spans:
            span = li.find('span')
            if span:
                job_areas.append(span.text)  # span 태그의 텍스트 값을 리스트에 추가

job_areas = {}
spans = ul_tag.find_all('li')
for idx, li in enumerate(spans, start=1):  
    span = li.find('span')
    if span:
        category_id = f"{idx:02}"
        job_areas[category_id] = span.text.strip()

print(job_areas) 

# 카테고리 이름만 리스트로 추출
category_list = list(job_areas.values())

# JavaScript 배열 형태로 저장
js_array = f"const categories = {json.dumps(category_list, ensure_ascii=False)};"

# 결과를 JS 파일로 저장
with open('categories.js', 'w', encoding='utf-8') as js_file:
    js_file.write(js_array)

print("[INFO] JavaScript 파일 'categories.js'가 성공적으로 생성되었습니다.")


#####################################자격증종목명 & 고유id 스크래핑 #################################################
# 기관명 매핑
org_names = {
    'N001': '한국산업인력공단',
    'P200': '대한상공회의소',
    'R139': '영화진흥위원회',
    'N004': '한국광해광업공단',
    'P317': '한국데이터산업진흥원',
    'R121': '한국디자인진흥원',
    'N003': '한국방송통신전파진흥원',
    'R020': '한국원자력안전기술원',
    'N002': '한국콘텐츠진흥원',
}

###################1. 한국산업인력공단 ########################
certifications = []  

# 한국산업인력공단 시행종목 (01~26)
for idx in range(1, 27):
    obligFldCd = str(format(idx, '02'))  # 1 -> '01', 2 -> '02', ...
    url = f'https://www.q-net.or.kr/crf005.do?id=crf00501s01&gSite=Q&gId=&div=1&obligFldCd={obligFldCd}'
    
    # 해당 URL에서 HTML 받아오기
    response = urlopen(url)
    soup = BeautifulSoup(response, 'html.parser')
    
    # 분야 ID 추출
    jmCd_ids = []  # 분야 ID를 저장할 리스트
    for a_tag in soup.find_all('a', onclick=True):
        match = re.search(r"step3BunRyu\('([0-9\-]+)',", a_tag['onclick'])
        if match:
            jmCd_id = match.group(1)  # 분야 ID
            jmCd_ids.append(jmCd_id)  # 리스트에 추가
    
    # 자격증명과 자격증종목 ID 추출
    for jmCd in jmCd_ids:
        detail_url = f'https://www.q-net.or.kr/crf005.do?id=crf00501p02&gSite=Q&gId=&jmCd={jmCd}&examInstiCd='
        response = urlopen(detail_url)
        soup = BeautifulSoup(response, 'html.parser')
        
        # 자격증명과 ID 추출
        for a_tag in soup.find_all('a', onclick=True):
            match = re.search(r"jmDetail\('(\d+)',\s*'([^']+)'\);", a_tag['onclick'])
            if match:
                cert_id = match.group(1)  # 자격증 ID
                cert_name = match.group(2)  # 자격증명

                # 카테고리 ID -> 명칭 변환
                category_name = job_areas.get(obligFldCd, "알 수 없는 카테고리")
                
                # 기관명, 카테고리 명칭, 자격증 ID, 자격증명 추가
                certifications.append((cert_id, cert_name, org_names['N001'] , category_name))


print(certifications)

#####################2. 타기관 ##########################

# 기관 코드와 카테고리, 기관명 매핑
org_categories = {
    'P200': ['10', '02'],  
    'R139': ['08'],  
    'N004': ['15'], 
    'P317': ['21'],  
    'R121': ['08'],  
    'N003': ['21'], 
    'R020': ['26'],  
    'N002': ['21'], 
}

# 분야 ID 저장
AllCertifications = []  # 모든 기관의 분야 ID 저장
CertificationDetails = []  # 모든 기관의 자격증 ID, 이름, 기관명 저장

# 모든 org와 카테고리를 조합하여 반복
for org, categories in org_categories.items():
    for category in categories:
        # 분야 ID 요청 URL
        url = f'https://www.q-net.or.kr/crf005.do?id=crf00501s03&gSite=Q&gId=&obligFldCd={category}&examInstiCd={org}'
        
        try:
            # 요청 및 HTML 파싱
            response = requests.get(url)
            soup = BeautifulSoup(response.content, 'html.parser')
            
            # 분야 ID (jmCd) 추출
            for input_tag in soup.find_all('input', {'name': 'jmcd'}):
                jmCd = input_tag.get('value')
                if jmCd:
                    AllCertifications.append((org, category, jmCd))  # 기관, 카테고리, jmCd 추가
        except Exception as e:
            print(f"Error fetching data for {org}, category {category}: {e}")

# 분야 ID를 통해 자격증 정보 추출
for org, category, jmCd in AllCertifications:
    # 자격증 요청 URL
    url = f'https://www.q-net.or.kr/crf005.do?id=crf00501s04&mdobligFldCd={jmCd}&examInstiCd={org}'
    
    try:
        # 요청 및 HTML 파싱
        response = requests.get(url)
        soup = BeautifulSoup(response.content, 'html.parser')
        
        # <a> 태그에서 자격증 ID와 이름 추출
        for a_tag in soup.find_all('a', onclick=True):
            match = re.search(r"jmDetail2\('(\d+)',\s*'([^']+)'\)", a_tag['onclick'])
            if match:
                cert_id = match.group(1)  # 자격증 ID
                cert_name = match.group(2)  # 자격증 이름

                 # 카테고리 ID -> 명칭 변환
                category = str(category).zfill(2)  # 항상 "01" 형식 유지
                category_name = job_areas.get(category, "알 수 없는 카테고리")  # 명칭 변환
                org_name = org_names.get(org, '알 수 없는 기관')  # 기관명 가져오기

                CertificationDetails.append(( cert_id, cert_name, org_name, category_name))  # 기관명, 카테고리, ID, 이름 저장
    except Exception as e:
        print(f"Error fetching data for jmCd {jmCd} in {org}, category {category_name}: {e}")


print("\n타기관 분야 ID 목록 (기관, 카테고리, jmCd):\n")
print(AllCertifications)

print("\n타기관 자격증 목록 (자격증 ID, 자격증 이름, 기관명, 카테고리):\n")
print(CertificationDetails)

####################################합격률&응시료  스크래핑#################################################

# 결과 리스트
certification_with_rates_and_fees = []

# 각 자격증 ID에 대해 합격률 및 응시료 정보 크롤링
for cert_id, cert_name, institution, category in CertificationDetails:
    ### 1. 합격률 정보 크롤링 ###
    url_rates = f'https://www.q-net.or.kr/crf005.do?id=crf00503s01&gSite=Q&gId=&jmCd={cert_id}'
    response_rates = urlopen(url_rates)
    soup_rates = BeautifulSoup(response_rates, 'html.parser')
    
    # 합격률 초기값
    written_pass_rate = None  # 필기 합격률
    practical_pass_rate = None  # 실기 합격률
    
    # "소 계" 데이터를 포함한 <tr> 태그 찾기
    for tr_tag in soup_rates.find_all('tr', class_='typeTotal'):
        td_tags = tr_tag.find_all('td')
        if len(td_tags) >= 6:  # 데이터가 충분한 경우
            written_pass_rate = td_tags[3].text.strip() if td_tags[3].text.strip() else None
            practical_pass_rate = td_tags[6].text.strip() if len(td_tags) > 6 and td_tags[6].text.strip() else None
            break  # 첫 번째 "소 계"만 처리
    
    ### 2. 응시료 정보 크롤링 ###
    url_fees = f"https://www.q-net.or.kr/crf005.do?id=crf00503s02&gSite=Q&gId=&jmCd={cert_id}"
    response_fees = requests.get(url_fees)
    soup_fees = BeautifulSoup(response_fees.text, 'html.parser')
    
    # 응시료 초기값
    written_fee = None  # 필기 응시료
    practical_fee = None  # 실기 응시료

    # 필기 및 실기 응시료가 포함된 <td> 태그 찾기
    fee_tags = soup_fees.find_all('td', {'scope': 'col'})
    if len(fee_tags) >= 1:  # 필기 응시료 존재 여부 확인
        written_fee = fee_tags[0].text.strip() if fee_tags[0].text.strip() else None
    if len(fee_tags) >= 2:  # 실기 응시료 존재 여부 확인
        practical_fee = fee_tags[1].text.strip() if fee_tags[1].text.strip() else None
    
    # 기존 데이터에 합격률 및 응시료 추가
    certification_with_rates_and_fees.append([
        cert_id, cert_name ,institution, category,
        practical_pass_rate,written_pass_rate,
        practical_fee,written_fee
    ])

# 최종 결과 출력
for cert in certification_with_rates_and_fees:
    print(f"자격증 ID: {cert[0]}, 자격증 이름: {cert[1]}, 기관명: {cert[2]}, 카테고리: {cert[3]}, "
          f"실기 합격률: {cert[4]}, 필기 합격률: {cert[5]}, 실기 응시료: {cert[6]}, 필기 응시료: {cert[7]}")

print(f"데이터 출력: {certification_with_rates_and_fees}")

####################################엑셀에 데이터 저장#################################################
###산업인력공단####
excel_filename = "certificationInfo.xlsx"
wb = Workbook()  # 새 워크북 생성
ws = wb.active
ws.title = "CertificationInfo"

# 헤더 추가
header = ['자격증 ID', '자격증 이름', '기관명', '카테고리', '실기 합격률', '필기 합격률', '실기 응시료', '필기 응시료']
ws.append(header)

# 첫 번째 데이터 추가
for row in certification_with_rates_and_fees:
    ws.append(row)

# 엑셀 저장
wb.save(excel_filename)
print(f"[INFO] '{excel_filename}' 파일 생성 및 첫 번째 데이터 저장 완료.")

# 기존 엑셀 파일 열기 및 두 번째 데이터 삽입
wb = load_workbook(excel_filename)
ws = wb.active  # 기존 워크시트 선택

# 두 번째 데이터 추가
for row in certification_with_rates_and_fees:
    ws.append(row)

# 엑셀 저장
wb.save(excel_filename)
print(f"[INFO] '{excel_filename}' 파일에 두 번째 데이터 추가 저장 완료.")


# 크롬 드라이버 종료를 생략하여 창을 유지하도록 변경
print("[INFO] 크롬 드라이버를 종료하지 않고 유지합니다. 수동으로 종료할 때까지 유지됩니다.")
time.sleep(60*5)  # 5분간 대기

# 크롬 드라이버 종료를 원할 때 직접 종료하려면 아래 코드로 종료할 수 있습니다:
# driver.quit()
